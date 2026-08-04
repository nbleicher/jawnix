from __future__ import annotations

import csv
import hashlib
import json
import sqlite3
import uuid
from collections.abc import Iterable
from datetime import date, datetime, time, timezone
from pathlib import Path

from sqlalchemy import select
from sqlalchemy.orm import Session

from jawnix.activity import record_activity
from jawnix.models import (
    Agency,
    Agent,
    CustomerProfile,
    DistributionEvent,
    Lead,
    LeadCorrectionEvent,
    LeadSource,
    ListingObservation,
    LeadRequest,
    MigrationAudit,
    QuarantinedRow,
    RequestStatus,
)
from jawnix.states import US_STATES, normalize_phone


def file_sha256(path: Path) -> str:
    digest = hashlib.sha256()
    with path.open("rb") as stream:
        for block in iter(lambda: stream.read(4 * 1024 * 1024), b""):
            digest.update(block)
    return digest.hexdigest()


def require_checksum(path: Path, expected: str | None) -> str:
    actual = file_sha256(path)
    if expected and actual.lower() != expected.lower():
        raise ValueError(f"Checksum mismatch for {path}: expected {expected}, got {actual}")
    return actual


def chunks(rows: Iterable[dict], size: int = 10_000):
    batch: list[dict] = []
    for row in rows:
        batch.append(row)
        if len(batch) >= size:
            yield batch
            batch = []
    if batch:
        yield batch


def import_agent_config(session: Session, path: Path) -> dict:
    config = json.loads(path.read_text(encoding="utf-8"))
    agents = [str(value).strip() for value in config.get("agents", []) if str(value).strip()]
    agent_states = config.get("agent_states") or {}
    for slug in agent_states:
        if slug not in agents:
            agents.append(slug)
    invalid = sorted(
        {
            str(state).strip().upper()
            for values in agent_states.values()
            for state in values
            if str(state).strip().upper() not in US_STATES
        }
    )
    if invalid:
        raise ValueError(
            "Invalid state codes in agent config require explicit correction before import: " + ", ".join(invalid)
        )

    agencies: dict[str, Agency] = {}
    for slug, members in (config.get("agencies") or {}).items():
        agency = session.scalar(select(Agency).where(Agency.slug == slug))
        if agency is None:
            agency = Agency(slug=slug, name=slug.replace("-", " ").title())
            session.add(agency)
            session.flush()
        agencies[slug] = agency
        for member in members:
            if member not in agents:
                agents.append(member)

    agency_by_agent = {
        member: agencies[agency_slug]
        for agency_slug, members in (config.get("agencies") or {}).items()
        for member in members
    }
    imported = 0
    for slug in agents:
        agent = session.scalar(select(Agent).where(Agent.slug == slug))
        if agent is None:
            agent = Agent(slug=slug, name=slug.replace("-", " ").title())
            session.add(agent)
            imported += 1
        agency = agency_by_agent.get(slug)
        agent.agency_id = agency.id if agency else None
        agent.active = True
    session.flush()
    return {"agents": len(agents), "created": imported, "agencies": len(agencies)}


def _upsert_lead_batch(session: Session, batch: list[dict], manifest_wins: bool) -> dict[str, Lead]:
    by_phone: dict[str, dict] = {}
    for row in batch:
        by_phone.setdefault(row["phone"], row)
    existing = {
        lead.phone: lead
        for lead in session.scalars(select(Lead).where(Lead.phone.in_(list(by_phone))))
    }
    for phone, row in by_phone.items():
        lead = existing.get(phone)
        if lead is None:
            title = row.get("title", "")
            state = row["state"]
            lead = Lead(
                phone=phone,
                title=title,
                state=state,
                # Seed the Legacy Listing Snapshot on first import so later
                # corrections have somewhere to fall back when no Current
                # Listing exists yet (mirrors alembic 0004 / sync behavior).
                legacy_title=title,
                legacy_state=state,
                source_flow=row.get("source_flow", ""),
                first_seen_at=row.get("first_seen_at") or datetime.now(timezone.utc),
                last_distributed_at=row.get("last_distributed_at"),
            )
            session.add(lead)
            existing[phone] = lead
        elif manifest_wins:
            if row.get("title"):
                lead.title = row["title"]
            lead.state = row["state"]
            if not lead.legacy_title and not lead.legacy_state:
                lead.legacy_title = lead.title
                lead.legacy_state = lead.state
            if row.get("last_distributed_at") and (
                lead.last_distributed_at is None or row["last_distributed_at"] > lead.last_distributed_at
            ):
                lead.last_distributed_at = row["last_distributed_at"]
    session.flush()
    return existing


def _record_sources(
    session: Session,
    leads: dict[str, Lead],
    rows: list[dict],
    source: str,
) -> None:
    by_phone = {row["phone"]: row for row in rows}
    keys = list(by_phone)
    existing = {
        item.source_key: item
        for item in session.scalars(
            select(LeadSource).where(LeadSource.source == source, LeadSource.source_key.in_(keys))
        )
    }
    now = datetime.now(timezone.utc)
    for phone, row in by_phone.items():
        item = existing.get(phone)
        metadata = {"flow": row.get("source_flow", ""), "row": row.get("row_number")}
        if item is None:
            session.add(
                LeadSource(
                    lead_id=leads[phone].id,
                    source=source,
                    source_key=phone,
                    metadata_json=metadata,
                    last_seen_at=now,
                )
            )
        else:
            item.last_seen_at = now
            item.metadata_json = metadata


def _is_postgresql(session: Session) -> bool:
    return session.get_bind().dialect.name == "postgresql"


def _postgres_driver_connection(session: Session):
    return session.connection().connection.driver_connection


def _postgres_prepare_manifest_staging(session: Session) -> None:
    connection = _postgres_driver_connection(session)
    with connection.cursor() as cursor:
        cursor.execute("SET LOCAL synchronous_commit = off")
        cursor.execute("SET LOCAL work_mem = '256MB'")
        cursor.execute(
            """
            CREATE TEMP TABLE manifest_import_stage (
                phone text NOT NULL,
                title text NOT NULL,
                state text NOT NULL,
                source_flow text NOT NULL,
                first_seen_at timestamptz,
                last_distributed_at timestamptz,
                agent_slug text,
                row_number bigint NOT NULL
            ) ON COMMIT DROP
            """
        )
        cursor.execute(
            """
            CREATE TEMP TABLE migration_quarantine_stage (
                source_path text NOT NULL,
                row_number bigint NOT NULL,
                reason text NOT NULL,
                raw_data jsonb NOT NULL
            ) ON COMMIT DROP
            """
        )


def _postgres_copy_manifest_rows(
    session: Session,
    valid: list[dict],
    invalid: list[tuple[str, int, str, str]],
) -> None:
    connection = _postgres_driver_connection(session)
    with connection.cursor() as cursor:
        if valid:
            with cursor.copy(
                """
                COPY manifest_import_stage (
                    phone, title, state, source_flow, first_seen_at,
                    last_distributed_at, agent_slug, row_number
                ) FROM STDIN
                """
            ) as copy:
                for row in valid:
                    copy.write_row(
                        (
                            row["phone"],
                            row["title"],
                            row["state"],
                            row["source_flow"],
                            row["first_seen_at"],
                            row["last_distributed_at"],
                            row["agent_slug"],
                            row["row_number"],
                        )
                    )
        if invalid:
            with cursor.copy(
                """
                COPY migration_quarantine_stage (
                    source_path, row_number, reason, raw_data
                ) FROM STDIN
                """
            ) as copy:
                for row in invalid:
                    copy.write_row(row)


def _postgres_finish_manifest(
    session: Session,
    path: Path,
    checksum: str,
    source_rows: int,
    imported: int,
    quarantined: int,
) -> dict:
    source = f"manifest:{checksum[:16]}"
    connection = _postgres_driver_connection(session)
    with connection.cursor() as cursor:
        cursor.execute("CREATE INDEX manifest_import_stage_phone_idx ON manifest_import_stage (phone)")
        cursor.execute(
            """
            INSERT INTO lead_inventory (
                phone, title, state, legacy_title, legacy_state,
                first_seen_at, source_flow, last_distributed_at
            )
            SELECT
                phone,
                COALESCE(MAX(NULLIF(title, '')), ''),
                MAX(state),
                COALESCE(MAX(NULLIF(title, '')), ''),
                MAX(state),
                COALESCE(MIN(first_seen_at), now()),
                COALESCE(MAX(NULLIF(source_flow, '')), 'manifest'),
                MAX(last_distributed_at)
            FROM manifest_import_stage
            GROUP BY phone
            ON CONFLICT (phone) DO UPDATE SET
                title = COALESCE(NULLIF(EXCLUDED.title, ''), lead_inventory.title),
                state = EXCLUDED.state,
                legacy_title = CASE
                    WHEN lead_inventory.legacy_title = ''
                         AND lead_inventory.legacy_state = ''
                    THEN COALESCE(NULLIF(EXCLUDED.title, ''), lead_inventory.title)
                    ELSE lead_inventory.legacy_title
                END,
                legacy_state = CASE
                    WHEN lead_inventory.legacy_title = ''
                         AND lead_inventory.legacy_state = ''
                    THEN EXCLUDED.state
                    ELSE lead_inventory.legacy_state
                END,
                source_flow = COALESCE(NULLIF(EXCLUDED.source_flow, ''), lead_inventory.source_flow),
                first_seen_at = LEAST(lead_inventory.first_seen_at, EXCLUDED.first_seen_at),
                last_distributed_at = CASE
                    WHEN lead_inventory.last_distributed_at IS NULL THEN EXCLUDED.last_distributed_at
                    WHEN EXCLUDED.last_distributed_at IS NULL THEN lead_inventory.last_distributed_at
                    ELSE GREATEST(lead_inventory.last_distributed_at, EXCLUDED.last_distributed_at)
                END
            """
        )
        cursor.execute(
            """
            INSERT INTO lead_sources (
                lead_id, source, source_key, metadata_json, first_seen_at, last_seen_at
            )
            SELECT
                inventory.id,
                %s,
                stage.phone,
                jsonb_build_object(
                    'flow', COALESCE(MAX(stage.source_flow), ''),
                    'row', MIN(stage.row_number)
                ),
                now(),
                now()
            FROM manifest_import_stage AS stage
            JOIN lead_inventory AS inventory ON inventory.phone = stage.phone
            GROUP BY inventory.id, stage.phone
            ON CONFLICT (source, source_key) DO UPDATE SET
                lead_id = EXCLUDED.lead_id,
                metadata_json = EXCLUDED.metadata_json,
                last_seen_at = EXCLUDED.last_seen_at
            """,
            (source,),
        )
        cursor.execute(
            """
            INSERT INTO distribution_events (
                lead_id, agent_id, request_id, delivered_at, source
            )
            SELECT DISTINCT
                inventory.id,
                agent.id,
                NULL::uuid,
                stage.last_distributed_at,
                CASE WHEN agent.id IS NULL
                    THEN 'legacy_unknown_recipient'
                    ELSE 'manifest'
                END
            FROM manifest_import_stage AS stage
            JOIN lead_inventory AS inventory ON inventory.phone = stage.phone
            LEFT JOIN agents AS agent ON agent.slug = stage.agent_slug
            WHERE stage.last_distributed_at IS NOT NULL
              AND NULLIF(stage.agent_slug, '') IS NOT NULL
            ON CONFLICT DO NOTHING
            """
        )
        cursor.execute(
            """
            INSERT INTO quarantined_rows (
                source_path, row_number, reason, raw_data, created_at
            )
            SELECT source_path, row_number, reason, raw_data, now()
            FROM migration_quarantine_stage
            """
        )
        cursor.execute(
            """
            INSERT INTO migration_audits (
                source_path, checksum, source_rows, imported_rows,
                quarantined_rows, completed_at
            ) VALUES (%s, %s, %s, %s, %s, now())
            """,
            (str(path), checksum, source_rows, imported, quarantined),
        )
    return {
        "sourceRows": source_rows,
        "imported": imported,
        "quarantined": quarantined,
        "checksum": checksum,
    }


def _import_manifest_postgresql(
    session: Session,
    path: Path,
    checksum: str,
) -> dict:
    _postgres_prepare_manifest_staging(session)
    source_rows = imported = quarantined = 0
    with path.open(newline="", encoding="utf-8", errors="replace") as stream:
        reader = csv.DictReader(stream)
        required = {"phone", "title", "state", "first_seen", "flow", "agent", "date_distributed"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise ValueError("Manifest header does not match the expected schema.")
        for batch in chunks(reader, size=50_000):
            valid: list[dict] = []
            invalid: list[tuple[str, int, str, str]] = []
            for row in batch:
                source_rows += 1
                row_number = source_rows + 1
                phone = normalize_phone(row.get("phone"))
                state = str(row.get("state") or "").strip().upper()
                if not phone or state not in US_STATES:
                    invalid.append(
                        (
                            str(path),
                            row_number,
                            "invalid phone or state",
                            json.dumps(dict(row), ensure_ascii=False),
                        )
                    )
                    quarantined += 1
                    continue
                first_seen = None
                if row.get("first_seen"):
                    try:
                        first_seen = datetime.combine(
                            date.fromisoformat(row["first_seen"]), time.min, tzinfo=timezone.utc
                        )
                    except ValueError:
                        pass
                delivered_at = None
                if row.get("date_distributed"):
                    try:
                        delivered_at = datetime.combine(
                            date.fromisoformat(row["date_distributed"]), time.min, tzinfo=timezone.utc
                        )
                    except ValueError:
                        pass
                valid.append(
                    {
                        "phone": phone,
                        "title": str(row.get("title") or ""),
                        "state": state,
                        "source_flow": str(row.get("flow") or "manifest"),
                        "first_seen_at": first_seen,
                        "last_distributed_at": delivered_at,
                        "agent_slug": str(row.get("agent") or "").strip(),
                        "row_number": row_number,
                    }
                )
            _postgres_copy_manifest_rows(session, valid, invalid)
            imported += len(valid)
    return _postgres_finish_manifest(session, path, checksum, source_rows, imported, quarantined)


def import_manifest(session: Session, path: Path, expected_checksum: str | None = None) -> dict:
    checksum = require_checksum(path, expected_checksum)
    existing_audit = session.scalar(
        select(MigrationAudit).where(MigrationAudit.source_path == str(path), MigrationAudit.checksum == checksum)
    )
    if existing_audit:
        return {"skipped": True, "sourceRows": existing_audit.source_rows, "imported": existing_audit.imported_rows}
    if _is_postgresql(session):
        return _import_manifest_postgresql(session, path, checksum)

    agents = {agent.slug: agent for agent in session.scalars(select(Agent))}
    source_rows = imported = quarantined = 0
    with path.open(newline="", encoding="utf-8", errors="replace") as stream:
        reader = csv.DictReader(stream)
        required = {"phone", "title", "state", "first_seen", "flow", "agent", "date_distributed"}
        if not required.issubset(set(reader.fieldnames or [])):
            raise ValueError("Manifest header does not match the expected schema.")
        for batch in chunks(reader):
            valid: list[dict] = []
            legacy: list[tuple[str, str, datetime]] = []
            for offset, row in enumerate(batch, start=source_rows + 2):
                source_rows += 1
                phone = normalize_phone(row.get("phone"))
                state = str(row.get("state") or "").strip().upper()
                if not phone or state not in US_STATES:
                    session.add(
                        QuarantinedRow(
                            source_path=str(path),
                            row_number=offset,
                            reason="invalid phone or state",
                            raw_data=dict(row),
                        )
                    )
                    quarantined += 1
                    continue
                first_seen = None
                if row.get("first_seen"):
                    try:
                        first_seen = datetime.combine(date.fromisoformat(row["first_seen"]), time.min, tzinfo=timezone.utc)
                    except ValueError:
                        pass
                delivered_at = None
                if row.get("date_distributed"):
                    try:
                        delivered_at = datetime.combine(date.fromisoformat(row["date_distributed"]), time.min, tzinfo=timezone.utc)
                    except ValueError:
                        pass
                valid.append(
                    {
                        "phone": phone,
                        "title": str(row.get("title") or ""),
                        "state": state,
                        "source_flow": str(row.get("flow") or "manifest"),
                        "first_seen_at": first_seen,
                        "last_distributed_at": delivered_at,
                        "row_number": offset,
                    }
                )
                if delivered_at and row.get("agent"):
                    legacy.append((phone, str(row["agent"]).strip(), delivered_at))
            leads = _upsert_lead_batch(session, valid, manifest_wins=True)
            _record_sources(session, leads, valid, f"manifest:{checksum[:16]}")
            legacy_lead_ids = [leads[phone].id for phone, _slug, _date in legacy]
            existing_events = {
                (event.lead_id, event.agent_id, event.delivered_at, event.source)
                for event in session.scalars(
                    select(DistributionEvent).where(
                        DistributionEvent.lead_id.in_(legacy_lead_ids),
                        DistributionEvent.source.in_(["manifest", "legacy_unknown_recipient"]),
                    )
                )
            } if legacy_lead_ids else set()
            for phone, agent_slug, delivered_at in legacy:
                agent = agents.get(agent_slug)
                source = "manifest" if agent else "legacy_unknown_recipient"
                key = (leads[phone].id, agent.id if agent else None, delivered_at, source)
                if key not in existing_events:
                    session.add(
                        DistributionEvent(
                            lead_id=leads[phone].id,
                            agent_id=agent.id if agent else None,
                            delivered_at=delivered_at,
                            source=source,
                        )
                    )
                    existing_events.add(key)
            imported += len(valid)
            session.flush()
    session.add(
        MigrationAudit(
            source_path=str(path),
            checksum=checksum,
            source_rows=source_rows,
            imported_rows=imported,
            quarantined_rows=quarantined,
        )
    )
    session.flush()
    return {"sourceRows": source_rows, "imported": imported, "quarantined": quarantined, "checksum": checksum}


def _postgres_prepare_scraper_staging(session: Session) -> None:
    connection = _postgres_driver_connection(session)
    with connection.cursor() as cursor:
        cursor.execute("SET LOCAL synchronous_commit = off")
        cursor.execute("SET LOCAL work_mem = '256MB'")
        cursor.execute(
            """
            CREATE TEMP TABLE scraper_import_stage (
                phone text NOT NULL,
                title text NOT NULL,
                state text NOT NULL,
                source_flow text NOT NULL
            ) ON COMMIT DROP
            """
        )
        cursor.execute(
            """
            CREATE TEMP TABLE migration_quarantine_stage (
                source_path text NOT NULL,
                row_number bigint NOT NULL,
                reason text NOT NULL,
                raw_data jsonb NOT NULL
            ) ON COMMIT DROP
            """
        )


def _postgres_copy_scraper_rows(
    session: Session,
    valid: list[dict],
    invalid: list[tuple[str, int, str, str]],
) -> None:
    connection = _postgres_driver_connection(session)
    with connection.cursor() as cursor:
        if valid:
            with cursor.copy(
                "COPY scraper_import_stage (phone, title, state, source_flow) FROM STDIN"
            ) as copy:
                for row in valid:
                    copy.write_row((row["phone"], row["title"], row["state"], row["source_flow"]))
        if invalid:
            with cursor.copy(
                """
                COPY migration_quarantine_stage (
                    source_path, row_number, reason, raw_data
                ) FROM STDIN
                """
            ) as copy:
                for row in invalid:
                    copy.write_row(row)


def _postgres_finish_scraper(
    session: Session,
    path: Path,
    checksum: str,
    source_rows: int,
    quarantined: int,
) -> dict:
    source = f"scraper:{path.name}"[:80]
    connection = _postgres_driver_connection(session)
    with connection.cursor() as cursor:
        cursor.execute("CREATE INDEX scraper_import_stage_phone_idx ON scraper_import_stage (phone)")
        cursor.execute(
            """
            INSERT INTO lead_inventory (
                phone, title, state, first_seen_at, source_flow, last_distributed_at
            )
            SELECT
                phone,
                COALESCE(MAX(NULLIF(title, '')), ''),
                MAX(state),
                now(),
                COALESCE(MAX(NULLIF(source_flow, '')), 'scraper:unknown'),
                NULL
            FROM scraper_import_stage
            GROUP BY phone
            ON CONFLICT (phone) DO NOTHING
            """
        )
        imported = cursor.rowcount
        cursor.execute(
            """
            INSERT INTO lead_sources (
                lead_id, source, source_key, metadata_json, first_seen_at, last_seen_at
            )
            SELECT
                inventory.id,
                %s,
                stage.phone,
                jsonb_build_object('flow', COALESCE(MAX(stage.source_flow), '')),
                now(),
                now()
            FROM scraper_import_stage AS stage
            JOIN lead_inventory AS inventory ON inventory.phone = stage.phone
            GROUP BY inventory.id, stage.phone
            ON CONFLICT (source, source_key) DO UPDATE SET
                lead_id = EXCLUDED.lead_id,
                metadata_json = EXCLUDED.metadata_json,
                last_seen_at = EXCLUDED.last_seen_at
            """,
            (source,),
        )
        cursor.execute(
            """
            INSERT INTO quarantined_rows (
                source_path, row_number, reason, raw_data, created_at
            )
            SELECT source_path, row_number, reason, raw_data, now()
            FROM migration_quarantine_stage
            """
        )
        cursor.execute(
            """
            INSERT INTO migration_audits (
                source_path, checksum, source_rows, imported_rows,
                quarantined_rows, completed_at
            ) VALUES (%s, %s, %s, %s, %s, now())
            """,
            (str(path), checksum, source_rows, imported, quarantined),
        )
    return {
        "sourceRows": source_rows,
        "imported": imported,
        "quarantined": quarantined,
        "checksum": checksum,
    }


def _import_scraper_postgresql(
    session: Session,
    path: Path,
    checksum: str,
) -> dict:
    _postgres_prepare_scraper_staging(session)
    connection = sqlite3.connect(f"file:{path}?mode=ro&immutable=1", uri=True)
    connection.row_factory = sqlite3.Row
    source_rows = quarantined = 0
    query = """
        SELECT phone,
               MAX(NULLIF(TRIM(company), '')) AS company,
               MAX(NULLIF(TRIM(full_name), '')) AS full_name,
               MAX(NULLIF(TRIM(niche), '')) AS niche,
               MAX(NULLIF(UPPER(TRIM(state)), '')) AS state,
               MAX(NULLIF(TRIM(source), '')) AS source,
               COUNT(*) AS source_count
        FROM leads
        WHERE phone IS NOT NULL AND TRIM(phone) != ''
        GROUP BY phone
    """
    try:
        cursor = connection.execute(query)
        while True:
            rows = cursor.fetchmany(50_000)
            if not rows:
                break
            valid: list[dict] = []
            invalid: list[tuple[str, int, str, str]] = []
            for row in rows:
                source_rows += int(row["source_count"] or 1)
                phone = normalize_phone(row["phone"])
                state = str(row["state"] or "").upper()
                if not phone or state not in US_STATES:
                    invalid.append(
                        (
                            str(path),
                            source_rows,
                            "invalid scraper phone or state",
                            json.dumps(
                                {"phone": row["phone"], "state": row["state"], "source": row["source"]},
                                ensure_ascii=False,
                            ),
                        )
                    )
                    quarantined += 1
                    continue
                valid.append(
                    {
                        "phone": phone,
                        "title": str(row["company"] or row["full_name"] or row["niche"] or ""),
                        "state": state,
                        "source_flow": f"scraper:{row['source'] or 'unknown'}",
                    }
                )
            _postgres_copy_scraper_rows(session, valid, invalid)
    finally:
        connection.close()
    return _postgres_finish_scraper(session, path, checksum, source_rows, quarantined)


def import_scraper_sqlite(session: Session, path: Path, expected_checksum: str | None = None) -> dict:
    checksum = require_checksum(path, expected_checksum)
    existing_audit = session.scalar(
        select(MigrationAudit).where(MigrationAudit.checksum == checksum)
    )
    if existing_audit:
        return {
            "skipped": True,
            "sourceRows": existing_audit.source_rows,
            "imported": existing_audit.imported_rows,
            "quarantined": existing_audit.quarantined_rows,
            "checksum": checksum,
        }
    connection = sqlite3.connect(f"file:{path}?mode=ro&immutable=1", uri=True)
    connection.row_factory = sqlite3.Row
    source_rows = imported = quarantined = observations = 0
    latest_valid: dict[str, dict] = {}
    leads: dict[str, Lead] = {}
    try:
        columns = {
            str(row["name"])
            for row in connection.execute("PRAGMA table_info(leads)")
        }
        required = {"phone", "company", "full_name", "niche", "state", "source"}
        missing = sorted(required - columns)
        if missing:
            raise ValueError(
                "Google Maps Scraper Dataset is missing column(s): "
                + ", ".join(missing)
            )
        observed_expression = (
            "created_at"
            if "created_at" in columns
            else "NULL"
        )
        cursor = connection.execute(
            f"""
            SELECT rowid AS source_row_number,
                   phone,
                   company,
                   full_name,
                   niche,
                   state,
                   source,
                   {observed_expression} AS observed_at
            FROM leads
            WHERE UPPER(COALESCE(source, '')) != 'NPPES'
            ORDER BY COALESCE({observed_expression}, ''), rowid
            """
        )
        for row in cursor:
            source_rows += 1
            row_number = int(row["source_row_number"])
            phone = normalize_phone(row["phone"])
            title = str(
                row["company"]
                or row["full_name"]
                or ""
            ).strip()
            state = str(row["state"] or "").strip().upper()
            source = str(row["source"] or "unknown").strip()
            niche = str(row["niche"] or "").strip()
            raw_observed_at = str(row["observed_at"] or "").strip()
            try:
                observed_at = datetime.fromisoformat(
                    raw_observed_at.replace("Z", "+00:00")
                )
                if observed_at.tzinfo is None:
                    observed_at = observed_at.replace(tzinfo=timezone.utc)
            except ValueError:
                observed_at = datetime(1970, 1, 1, tzinfo=timezone.utc)
                observed_at = observed_at.replace(
                    microsecond=row_number % 1_000_000
                )

            valid = bool(phone and title and state in US_STATES)
            lead = None
            if phone:
                lead = leads.get(phone)
                if lead is None:
                    lead = session.scalar(
                        select(Lead).where(Lead.phone == phone)
                    )
                    if lead is not None:
                        leads[phone] = lead
            if valid and lead is None:
                lead = Lead(
                    phone=phone,
                    title=title,
                    state=state,
                    source_flow=f"google_maps:{source}"[:80],
                )
                session.add(lead)
                session.flush()
                leads[phone] = lead
                imported += 1

            observation = ListingObservation(
                lead_id=lead.id if lead else None,
                dataset_checksum=checksum,
                row_number=row_number,
                normalized_phone=phone or "",
                title=title,
                state=state if state in US_STATES else "",
                source=source[:160],
                niche=niche[:160],
                valid=valid,
                observed_at=observed_at,
                raw_data={
                    "phone": row["phone"],
                    "company": row["company"],
                    "full_name": row["full_name"],
                    "niche": row["niche"],
                    "state": row["state"],
                    "source": row["source"],
                    "observed_at": row["observed_at"],
                },
            )
            session.add(observation)
            session.flush()
            observations += 1

            if not valid:
                missing_fields = []
                if not phone:
                    missing_fields.append("valid US phone")
                if not title:
                    missing_fields.append("title")
                if state not in US_STATES:
                    missing_fields.append("Google Maps business state")
                session.add(
                    QuarantinedRow(
                        source_path=str(path),
                        row_number=row_number,
                        reason="invalid Google Maps listing: "
                        + ", ".join(missing_fields),
                        raw_data=observation.raw_data,
                    )
                )
                quarantined += 1
                continue

            current = (
                session.get(
                    ListingObservation,
                    lead.current_listing_observation_id,
                )
                if lead.current_listing_observation_id is not None
                else None
            )
            current_observed_at = current.observed_at if current else None
            if (
                current_observed_at is not None
                and current_observed_at.tzinfo is None
            ):
                current_observed_at = current_observed_at.replace(
                    tzinfo=timezone.utc
                )
            observation_order = (
                observed_at,
                checksum,
                row_number,
            )
            current_order = (
                current_observed_at,
                current.dataset_checksum,
                current.row_number,
            ) if current and current_observed_at else None
            if current_order is None or observation_order > current_order:
                if (
                    current is None
                    and not lead.legacy_title
                    and not lead.source_flow.startswith("google_maps:")
                ):
                    lead.legacy_title = lead.title
                    lead.legacy_state = lead.state
                lead.source_flow = f"google_maps:{source}"[:80]
                lead.current_listing_observation_id = observation.id
                correction = (
                    session.get(
                        LeadCorrectionEvent,
                        lead.active_correction_id,
                    )
                    if lead.active_correction_id is not None
                    else None
                )
                if correction is None:
                    lead.title = title
                    lead.state = state
                elif (
                    correction.title != title
                    or correction.state != state
                ):
                    record_activity(
                        session,
                        action="listing_observation_conflicts_correction",
                        target_type="listing_observation",
                        target_id=observation.id,
                        actor_id="system:inventory_sync",
                        reason=(
                            "A newer Google Maps listing conflicts "
                            "with the active Lead Correction."
                        ),
                        details={
                            "before": {
                                "title": correction.title,
                                "state": correction.state,
                            },
                            "after": {
                                "title": title,
                                "state": state,
                            },
                            "leadId": lead.id,
                            "correctionId": str(correction.id),
                        },
                    )
            latest_valid[phone] = {
                "phone": phone,
                "title": title,
                "state": state,
                "source_flow": lead.source_flow,
            }
    finally:
        connection.close()
    if latest_valid:
        _record_sources(
            session,
            leads,
            list(latest_valid.values()),
            f"google_maps:{path.name}"[:80],
        )
    session.add(
        MigrationAudit(
            source_path=str(path),
            checksum=checksum,
            source_rows=source_rows,
            imported_rows=imported,
            quarantined_rows=quarantined,
        )
    )
    return {
        "sourceRows": source_rows,
        "imported": imported,
        "quarantined": quarantined,
        "observations": observations,
        "checksum": checksum,
    }


def import_supabase_jsonl(session: Session, directory: Path) -> dict:
    import gzip

    profiles_path = directory / "jawnix_profiles.jsonl.gz"
    requests_path = directory / "jawnix_lead_requests.jsonl.gz"
    profile_count = request_count = request_skipped_unmapped = 0
    if profiles_path.exists():
        with gzip.open(profiles_path, "rt", encoding="utf-8") as stream:
            for line in stream:
                row = json.loads(line)
                user_id = uuid.UUID(str(row["user_id"]))
                profile = session.get(CustomerProfile, user_id)
                if profile is None:
                    profile = CustomerProfile(user_id=user_id, email=str(row.get("email") or "").lower())
                    session.add(profile)
                profile.email = str(row.get("email") or profile.email).lower()
                profile.first_name = str(row.get("first_name") or "")
                profile.last_name = str(row.get("last_name") or "")
                profile.phone = str(row.get("phone") or "")
                profile.licensed_states = [s for s in row.get("licensed_states") or [] if s in US_STATES]
                profile_count += 1
    session.flush()
    if requests_path.exists():
        with gzip.open(requests_path, "rt", encoding="utf-8") as stream:
            for line in stream:
                row = json.loads(line)
                profile = session.get(CustomerProfile, uuid.UUID(str(row["user_id"])))
                if profile is None or profile.customer_id is None:
                    request_skipped_unmapped += 1
                    continue
                legacy_id = uuid.UUID(str(row["id"]))
                if session.get(LeadRequest, legacy_id):
                    continue
                status = RequestStatus.delivered.value if row.get("status") == "fulfilled" else RequestStatus.pending.value
                session.add(
                    LeadRequest(
                        id=legacy_id,
                        user_id=profile.user_id,
                        agent_id=profile.customer_id,
                        lead_count=int(row["lead_count"]),
                        state_mode="all_saved",
                        states_snapshot=row.get("states_snapshot") or profile.licensed_states,
                        delivery_email=profile.email,
                        status=status,
                        status_message="Imported from Supabase.",
                    )
                )
                request_count += 1
    return {
        "profiles": profile_count,
        "requests": request_count,
        "requestsSkippedUnmapped": request_skipped_unmapped,
    }


def _history_recipient(filename: str, agents: dict[str, Agent]) -> Agent | None:
    normalized = "-" + "".join(character.lower() if character.isalnum() else "-" for character in filename) + "-"
    matches = [agent for slug, agent in agents.items() if f"-{slug.lower()}-" in normalized]
    return matches[0] if len(matches) == 1 else None


def _history_phones(path: Path) -> Iterable[tuple[int, str | None]]:
    candidates = ["phone", "phone number", "mobile phone", "mobile"]
    if path.suffix.lower() == ".csv":
        with path.open(newline="", encoding="utf-8", errors="replace") as stream:
            reader = csv.reader(stream)
            header = next(reader, [])
            normalized = [str(value).strip().lower().replace("_", " ") for value in header]
            index = next((normalized.index(value) for value in candidates if value in normalized), None)
            if index is None:
                return
            for row_number, row in enumerate(reader, start=2):
                yield row_number, row[index] if index < len(row) else None
        return
    if path.suffix.lower() == ".xlsx":
        from openpyxl import load_workbook

        workbook = load_workbook(path, read_only=True, data_only=True)
        try:
            rows = workbook.active.iter_rows(values_only=True)
            header = next(rows, ())
            normalized = [str(value or "").strip().lower().replace("_", " ") for value in header]
            index = next((normalized.index(value) for value in candidates if value in normalized), None)
            if index is None:
                return
            for row_number, row in enumerate(rows, start=2):
                yield row_number, row[index] if index < len(row) else None
        finally:
            workbook.close()


def import_distribution_history(session: Session, directory: Path) -> dict:
    """Reconstruct identifiable recipient history without modifying source files."""
    agents = {agent.slug: agent for agent in session.scalars(select(Agent))}
    files_seen = files_imported = events_created = skipped_recipient = malformed = 0
    for path in sorted(directory.rglob("*")):
        if not path.is_file() or path.suffix.lower() not in {".csv", ".xlsx"}:
            continue
        files_seen += 1
        agent = _history_recipient(path.name, agents)
        try:
            delivered_at = datetime.combine(date.fromisoformat(path.name[:10]), time.min, tzinfo=timezone.utc)
        except ValueError:
            delivered_at = None
        if agent is None or delivered_at is None:
            skipped_recipient += 1
            continue
        relative = str(path.relative_to(directory))
        source = f"history:{hashlib.sha256(relative.encode()).hexdigest()[:12]}:{path.name}"[:80]
        for batch in chunks(({"row": number, "phone": value} for number, value in _history_phones(path))):
            valid: dict[str, int] = {}
            for row in batch:
                phone = normalize_phone(row["phone"])
                if phone:
                    valid.setdefault(phone, row["row"])
                else:
                    session.add(
                        QuarantinedRow(
                            source_path=str(path),
                            row_number=row["row"],
                            reason="invalid history phone",
                            raw_data={"phone": row["phone"]},
                        )
                    )
                    malformed += 1
            if not valid:
                continue
            leads = {lead.phone: lead for lead in session.scalars(select(Lead).where(Lead.phone.in_(valid)))}
            for phone in set(valid) - set(leads):
                session.add(
                    QuarantinedRow(
                        source_path=str(path),
                        row_number=valid[phone],
                        reason="history phone not present in inventory",
                        raw_data={"phone": phone},
                    )
                )
            existing = {
                event.lead_id
                for event in session.scalars(
                    select(DistributionEvent).where(
                        DistributionEvent.lead_id.in_([lead.id for lead in leads.values()]),
                        DistributionEvent.agent_id == agent.id,
                        DistributionEvent.delivered_at == delivered_at,
                        DistributionEvent.source == source,
                    )
                )
            }
            for lead in leads.values():
                if lead.id not in existing:
                    session.add(
                        DistributionEvent(
                            lead_id=lead.id,
                            agent_id=agent.id,
                            customer_name=agent.name,
                            agency_id=agent.agency_id,
                            agency_name=agent.agency.name if agent.agency else "",
                            phone=lead.phone,
                            title=lead.title,
                            state=lead.state,
                            listing_provenance={
                                "kind": "legacy",
                                "source": lead.source_flow,
                            },
                            source_kind="legacy",
                            delivered_at=delivered_at,
                            source=source,
                        )
                    )
                    events_created += 1
                if lead.last_distributed_at is None or delivered_at > lead.last_distributed_at:
                    lead.last_distributed_at = delivered_at
            session.flush()
        files_imported += 1
    return {
        "filesSeen": files_seen,
        "filesImported": files_imported,
        "eventsCreated": events_created,
        "skippedUnidentifiedRecipient": skipped_recipient,
        "malformedPhones": malformed,
    }
